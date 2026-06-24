"""
Management command: ingest interpretacji indywidualnych KIS z EUREKA.

Sposoby podania interpretacji:

  # 1) wprost po ID informacji
  python manage.py ingest_interpretacje --ids 604348,639490 --ulga IPBOX

  # 2) z eksportu wyszukiwarki EUREKA (CSV/XLSX)
  python manage.py ingest_interpretacje --csv eksport_eureka.xlsx --ulga BR

  # 3) wyszukiwanie po ID przepisów (PRZEPISY)
  python manage.py ingest_interpretacje --przepisy 35573,40951 --limit 50

  # 4) NAJPROŚCIEJ — ulga sama dobiera przepisy z config.PRZEPISY_BY_ULGA
  python manage.py ingest_interpretacje --ulga IPBOX --limit 50 --od-daty 2023-01-01

Wyszukiwanie filtruje serwerowo do kategorii „Interpretacja indywidualna" i statusu
„Aktualna" (kody z configu), opcjonalnie po dacie wydania (--od-daty/--do-daty).
"""

import re

from django.core.management.base import BaseCommand, CommandError

from config import (
    KIS_KATEGORIA_INTERPRETACJA_ID,
    KIS_STATUS_AKTUALNA_ID,
    PRZEPISY_BY_ULGA,
)
from ulgi.ingest_docs import ingest_interpretacja_to_stores
from ulgi.kis_client import search_interpretacje

_RX_PODGLAD = re.compile(r"/podglad/(\d+)")


def _scan_cells(rows) -> list[str]:
    ids: list[str] = []
    for row in rows:
        for cell in row:
            s = str(cell if cell is not None else "")
            m = _RX_PODGLAD.search(s)
            if m:
                ids.append(m.group(1))
            elif s.isdigit() and 4 <= len(s) <= 9:
                ids.append(s)
    return ids


def _ids_from_export(path: str) -> list[str]:
    """Wyciąga ID informacji z eksportu EUREKA (CSV/XLSX)."""
    if path.lower().endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True)
        ids = _scan_cells(wb.active.iter_rows(values_only=True))
    else:
        import csv

        with open(path, newline="", encoding="utf-8-sig") as f:
            sample = f.read(2048)
            f.seek(0)
            delim = ";" if sample.count(";") >= sample.count(",") else ","
            ids = _scan_cells(csv.reader(f, delimiter=delim))
    return list(dict.fromkeys(ids))  # unikalne, kolejność zachowana


class Command(BaseCommand):
    help = "Ingest interpretacji indywidualnych KIS (EUREKA) do PostgreSQL + OpenSearch."

    def add_arguments(self, parser):
        parser.add_argument("--ids", help="ID informacji po przecinku, np. 604348,639490")
        parser.add_argument(
            "--csv", help="eksport EUREKA (CSV/XLSX) — ID wyłuskiwane automatycznie"
        )
        parser.add_argument(
            "--przepisy",
            help="ID przepisów (PRZEPISY) po przecinku — wyszukiwanie w EUREKA",
        )
        parser.add_argument(
            "--limit", type=int, default=50, help="max interpretacji z wyszukiwania"
        )
        parser.add_argument("--od-daty", dest="od_daty", help="data wydania od (YYYY-MM-DD)")
        parser.add_argument("--do-daty", dest="do_daty", help="data wydania do (YYYY-MM-DD)")
        parser.add_argument(
            "--ulga",
            default="",
            help="tag ulgi: BR / IPBOX / PKUP. Bez --ids/--csv/--przepisy sam dobiera przepisy.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="tylko wypisz znalezione interpretacje, nie indeksuj",
        )

    def _search_ids(self, przepisy, opts) -> list[str]:
        self.stdout.write(
            f"Wyszukiwanie EUREKA: przepisy={przepisy}, "
            f"kategoria=Interpretacja indywidualna, status=Aktualna, "
            f"data {opts['od_daty'] or '—'}…{opts['do_daty'] or '—'}, limit={opts['limit']}..."
        )
        hits = search_interpretacje(
            przepisy,
            kategoria_id=KIS_KATEGORIA_INTERPRETACJA_ID,
            status_id=KIS_STATUS_AKTUALNA_ID,
            od_daty=opts["od_daty"],
            do_daty=opts["do_daty"],
            limit=opts["limit"],
        )
        self.stdout.write(self.style.SUCCESS(f"Znaleziono {len(hits)} interpretacji:"))
        for h in hits:
            self.stdout.write(f"  {h['data']}  id={h['id']:>7}  {h['sygnatura']}")
        return [h["id"] for h in hits]

    def _przepisy_for_search(self, opts) -> list | None:
        """Zwraca listę ID przepisów do wyszukania albo None (gdy nie szukamy)."""
        if opts["przepisy"]:
            return [x.strip() for x in opts["przepisy"].split(",") if x.strip()]
        # --ulga sam dobiera przepisy, o ile nie podano ID/CSV
        if opts["ulga"] and not opts["ids"] and not opts["csv"]:
            przepisy = PRZEPISY_BY_ULGA.get(opts["ulga"].upper())
            if not przepisy:
                raise CommandError(
                    f"Brak mapy przepisów dla ulgi '{opts['ulga']}'. "
                    f"Dostępne: {', '.join(PRZEPISY_BY_ULGA)}"
                )
            return przepisy
        return None

    def handle(self, *args, **opts):
        ids: list[str] = []
        if opts["ids"]:
            ids += [x.strip() for x in opts["ids"].split(",") if x.strip()]
        if opts["csv"]:
            ids += _ids_from_export(opts["csv"])

        przepisy = self._przepisy_for_search(opts)
        if przepisy:
            ids += self._search_ids(przepisy, opts)

        ids = list(dict.fromkeys(ids))
        if not ids:
            raise CommandError(
                "Podaj --ids, albo --csv, albo --przepisy, albo --ulga (BR/IPBOX/PKUP)."
            )

        if opts["dry_run"]:
            self.stdout.write(
                self.style.WARNING(f"--dry-run: {len(ids)} ID, pomijam indeksowanie.")
            )
            return

        ok_n = 0
        for info_id in ids:
            self.stdout.write(f"Ingest interpretacji {info_id}...")
            try:
                out = ingest_interpretacja_to_stores(info_id, ulga=opts["ulga"])
                self.stdout.write(
                    self.style.SUCCESS(
                        f"  {out['interpretacja']}: ok={out['ok']}, błędy={out['errors']}"
                    )
                )
                ok_n += 1
            except Exception as e:  # noqa: BLE001
                self.stderr.write(self.style.ERROR(f"  {info_id}: BŁĄD — {e}"))
        self.stdout.write(self.style.SUCCESS(f"Zaindeksowano {ok_n}/{len(ids)} interpretacji."))
